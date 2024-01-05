import zipfile
import pandas as pd
from openpyxl import load_workbook
import os


current_dir = os.path.dirname(os.path.realpath(__file__))
zip_filename = "victims.zip"
zip_path = os.path.join(current_dir, zip_filename)

if os.path.exists(zip_path):
    print('ZIP file exists!')
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        excel_filename = "Victims_Age_by_Offense_Category_2022.xlsx"
        if excel_filename in zip_ref.namelist():
            print(f'Excel file found: {excel_filename}')

                    # read excel file
            zip_ref.extract(excel_filename, current_dir)
            excel_path = os.path.join(current_dir, excel_filename)

            workbook = load_workbook(excel_path)
            sheet_name = workbook.sheetnames[0]
            df = pd.read_excel(excel_path, sheet_name)

                    # filter category 'Crimes Against Property'
            rows_to_filter = list(range(12, 25))
            df_filtered = df.iloc[rows_to_filter]

            df_filtered = df_filtered.iloc[:-1, 2:16 ]
            df_filtered.reset_index(drop=True, inplace=True)

                    # write CSV file
            csv_filename = "crimes_property_data.csv"
            csv_path = os.path.join(current_dir, csv_filename)
            df_filtered.to_csv(csv_path, index=False,header=False)
            
            print(f'Leaked data saved in: {csv_filename}')
            os.remove(excel_path)

        else:
            print(f'Excel file not found: {excel_filename}')
else:
    print('ZIP file does not exist')




